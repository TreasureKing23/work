import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from io import BytesIO

REGISTER_MAP ={
    "ExamCode": "EXAMID", "ExamYear": "PERIODID", "EXCID": "CANID", "ERN": "ERN",
    "FirstName": "FNAME", "Lastname": "SURNAME", "Middlename": "MNAME",
    "Schoolcode": "SCHOOLID", "SchoolName": "SCHOOLN", "Gender": "GENDER",
    "DateofBirth": "DOB", "Address": "ADDR1", "Address_2": "ADDR2",
    "ContactEmail": "EMAIL", "LMS_Account": "LMS", "PathNo": "PATH",
    "WardofState": "WARD", "Sector": "SECTOR", "Parish": "PARISH", "Region": "REGION",
    "Prefcode1": "PCODE1", "First_Preference": "PNAME1",
    "Prefcode2": "PCODE2", "Second_Preference": "PNAME2",
    "Prefcode3": "PCODE3", "Third_Preference": "PNAME3",
    "Prefcode4": "PCODE4", "Fourth_Preference": "PNAME4",
    "Prefcode5": "PCODE5", "Fifth_Preference": "PNAME5",
    "Prefcode6": "CCODE1", "Sixth_Preference": "CNAME1",
    "Prefcode7": "CCODE2", "Seventh_Preference": "CNAME2",
    "SRN": "SRN"
}

COUNT_MAP={
    "SchoolCode" : "SCHOOLID" , "SchoolName": "SCHOOLN", "Parish":"PARISH", "Region":"REGION",
    "Male":"MALE_CNT", "Female": "FEM_CNT", "Total": "TOTAL"
}

PAPERS ={
    "PEP6": [
        "ABILITY", "MATHEMATICS - PT", "LANGUAGE ARTS - PT",
        "MATHEMATICS - CBT", "SCIENCE - CBT", "SOCIAL STUDIES - CBT", "LANGUAGE ARTS - CBT"
    ],
    "PEP5": [
        "MATHEMATICS - PT", "SCIENCE - PT", "SOCIAL STUDIES - PT", "LANGUAGE ARTS - PT"
    ],
    "PEP4": [
        "MATHEMATICS - PT", "NUMERACY", "LANGUAGE ARTS - PT", "LITERACY"
    ]
}

def formatting(input_specification: BytesIO,
               data_workbook: BytesIO,
               exam_name: str,
               exam_period:str,
               mode: str) -> BytesIO:

    dw_data = pd.read_excel(data_workbook)
    wb = load_workbook(input_specification)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    if mode == "count":
        dw_data.columns = dw_data.columns.str.strip()
        for col in ("Male", "Female", "Total"):
            if col not in dw_data.columns:
                dw_data[col]= 0
            dw_data[col] = pd.to_numeric(dw_data[col], errors="coerce").fillna(0)

        group_cols = ["SchoolCode", "SchoolName", "Parish","Region"]
        dw_data = (dw_data.groupby(group_cols, as_index=False)[["Male","Female","Total"]].sum())

        if "Total" not in dw_data.columns or dw_data["Total"].eq(0).all():
            dw_data["Total"] = dw_data["Male"] + dw_data["Female"]

        append_rows=[]
        for _, data in dw_data.iterrows():
            new_row=[]
            for header in headers:
                if header == "EXAMID":
                    new_row.append(exam_name)
                elif header == "PERIODID":
                    new_row.append(exam_period)

                else:
                    data_key = next ((k for k, v in COUNT_MAP.items() if v == header), None)
                    if data_key and data_key in data:
                        val = data[data_key]

                        if header == "SCHOOLID":
                                val = str(val).zfill(5)

                        if header == "REGION" and isinstance(val, str):
                            if "." in val:
                                val = val.split(".", 1)[-1].strip()

                        elif header == "SCHOOLN" and isinstance(val, str):
                            if "[" in val:
                                val = val.split("[")[0].strip()

                        new_row.append(val)
                    else:
                        new_row.append("")
            append_rows.append(new_row)

        for r in append_rows:
            ws.append(r)
        alignment = Alignment(horizontal="left")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = alignment

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


    paper_name = PAPERS.get(exam_name, [])
    append_rows=[]
    for _, data in dw_data.iterrows():
        new_row=[]

        for header in headers:
            data_key = next ((k for k, v in REGISTER_MAP.items() if v == header), None)

            if data_key in data:
                val = data[data_key]

                if header == "SCHOOLID":
                    val = str(val).zfill(5)

                if header == "REGION" and isinstance(val, str):
                    if "." in val:
                        val = val.split(".", 1)[-1].strip()


                if header == "SCHOOLN" and isinstance(val, str):
                    if "[" in val:
                        val = val.split("[")[0].strip()

                # Clean values like "CALABAR HIGH -> [ST. ANDREW]"
                if header.startswith("PNAME") or header.startswith("CNAME"):
                    if isinstance(val, str) and "->" in val:
                        val = val.split("->")[0].strip()



                if header in ("PATH", "WARD"):
                    val = "Y" if pd.notna(val) and str(val).strip() else "N"
                new_row.append(val)

            elif header == "TEL_NUM":
                if pd.notna(data.get("MotherContact")) and str(data.get("MotherContact")).strip():
                    val = data.get("MotherContact", "")
                elif pd.notna(data.get("FatherContact")) and str(data.get("FatherContact")).strip():
                    val = data.get("FatherContact", "")
                elif pd.notna(data.get("GuardianContact")) and str(data.get("GuardianContact")).strip():
                    val = data.get("GuardianContact", "")
                else:
                    val = ""

                if isinstance(val, str):
                    val = val.strip()
                # Remove if not long enough
                if len(val) < 14:
                    val = ""
                # Remove if area code is not (876) or (658)
                elif not (val.startswith("(876)") or val.startswith("(658)")):
                    val = ""

                new_row.append(val)

            elif header == "GNAME":
                if pd.notna(data.get("MotherContact")) and str(data.get("MotherContact")).strip():
                    val = data.get("Mother", "")
                elif pd.notna(data.get("FatherContact")) and str(data.get("FatherContact")).strip():
                    val = data.get("Father", "")
                elif pd.notna(data.get("GuardianContact")) and str(data.get("GuardianContact")).strip():
                    val = data.get("Gaurdian", "")
                else:
                    if pd.notna(data.get("Mother")):
                        val = data.get("Mother", "")
                    elif pd.notna(data.get("Father")):
                        val = data.get("Father", "")
                    elif pd.notna(data.get("Guardian")):
                        val = data.get("Guardian", "")
                    else:
                        val = ""
                new_row.append(val)

            elif header.startswith("PAPER0") and header[-2]:
                idx = int(header[-2:]) - 1
                new_row.append(paper_name[idx] if idx < len(paper_name) else "")
            else:
                new_row.append(None)

        append_rows.append(new_row)

    for r in append_rows:
        ws.append(r)
    alignment = Alignment(horizontal="left")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output