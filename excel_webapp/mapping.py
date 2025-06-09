# mapping.py
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

FIELD_MAP = {
    "ExamCode": "EXAMID", "ExamYear": "PERIODID", "EXCID": "CANID", "ERN": "ERN",
    "FirstName": "FNAME", "Lastname": "SURNAME", "Middlename": "MNAME",
    "SchoolId": "SCHOOLID", "SchoolName": "SCHOOLN", "Gender": "GENDER",
    "DateofBirth": "DOB", "Address": "ADDR1", "Address_2": "ADDR2",
    "ContactEmail": "EMAIL", "LMS_Account": "LMS", "PathNo": "PATH",
    "WardofState": "WARD", "Sector": "SECTOR", "Parish": "PARISH", "Region": "REGION",
    "Prefcode1": "PCODE1", "First_Preference": "PNAME1",
    "Prefcode2": "PCODE2", "Second_Preference": "PNAME2",
    "Prefcode3": "PCODE3", "Third_Preference": "PNAME3",
    "Prefcode4": "PCODE4", "Fourth_Preference": "PNAME4",
    "Prefcode5": "PCODE5", "Fifth_Preference": "PNAME5",
    "Prefcode6": "CCODE1", "Sixth_Preference": "CNAME1",
    "Prefcode7": "CCODE2", "Seventh_Preference": "CNAME2",
    "SRN": "SRN"
}

PAPERS = {
    "PEP6": [
        "Ability", "Mathematics PT", "Language Arts PT",
        "Mathematics CBT", "Science CBT", "Social Studies CBT", "Language Arts CBT"
    ],
    "PEP5": [
        "Mathematics PT", "Science PT", "Social Studies PT", "Language PT"
    ],
    "PEP4": [
        "Mathematics PT", "Numeracy", "Language PT", "Literacy"
    ]
}


def format_workbook(template_file: BytesIO,
                    data_file: BytesIO,
                    exam: str) -> BytesIO:
    """
    Runs the mapping/formatting logic and returns an in-memory xlsx stream.
    """

    # --- Load workbooks
    df_src = pd.read_excel(data_file)
    wb = load_workbook(template_file)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    paper_names = PAPERS.get(exam, [])

    rows_to_append = []
    for _, src in df_src.iterrows():
        new_row = []

        for header in headers:
            src_key = next((k for k, v in FIELD_MAP.items() if v == header), None)

            if src_key in src:
                val = src[src_key]
                if header in ("PATH", "WARD"):
                    val = "Y" if pd.notna(val) and str(val).strip() else "N"
                new_row.append(val)

            elif header == "TEL_NUM":
                if pd.notna(src.get("MotherContact")) and str(src.get("MotherContact")).strip():
                    val = src.get("MotherContact", "")
                elif pd.notna(src.get("FatherContact")) and str(src.get("FatherContact")).strip():
                    val = src.get("FatherContact", "")
                elif pd.notna(src.get("GuardianContact")) and str(src.get("GuardianContact")).strip():
                    val = src.get("GuardianContact", "")
                else:
                    val = ""
                new_row.append(val)
            
            elif header == "GNAME":
                if pd.notna(src.get("MotherContact")) and str(src.get("MotherContact")).strip():
                    val = src.get("Mother", "")
                elif pd.notna(src.get("FatherContact")) and str(src.get("FatherContact")).strip():
                    val = src.get("Father", "")
                elif pd.notna(src.get("GuardianContact")) and str(src.get("GuardianContact")).strip():
                    val = src.get("Guardian", "")
                else:
                    val = ""
                new_row.append(val)


            elif header.startswith("PAPER0") and header[-2:].isdigit():
                idx = int(header[-2:]) - 1
                new_row.append(paper_names[idx] if idx < len(paper_names) else "")
            else:
                new_row.append(None)

        rows_to_append.append(new_row)

    # append rows to template
    for r in rows_to_append:
        ws.append(r)

    # Save workbook to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
