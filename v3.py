import pandas as pd
from openpyxl import load_workbook
from tkinter import Tk, filedialog, messagebox
import os

# Hide the root window
root = Tk()
root.withdraw()

# Ask user to select the Input and Data files
messagebox.showinfo("Select Files", "Select the INPUT Excel file (Workbook1)")
input_path = filedialog.askopenfilename(title="Select Input Workbook")

messagebox.showinfo("Select Files", "Select the DATA Excel file (Workbook2)")
data_path = filedialog.askopenfilename(title="Select Data Workbook")

# Ask for output file location
output_path = filedialog.asksaveasfilename(
    defaultextension=".xlsx",
    title="Save Updated Workbook As",
    filetypes=[("Excel Files", "*.xlsx")]
)

if not input_path or not data_path or not output_path:
    messagebox.showerror("Cancelled", "Operation cancelled. All files must be selected.")
    exit()

# Field mapping from Workbook2 → Workbook1
field_map = {
    "ExamCode": "EXAMID",
    "ExamYear": "PERIODID",
    "EXCID": "CANID",
    "ERN": "ERN",
    "FirstName": "FNAME",
    "Lastname": "SURNAME",
    "Middlename": "MNAME",
    "SchoolId": "SCHOOLID",
    "SchoolName": "SCHOOLN",
    "Gender": "GENDER",
    "DateofBirth": "DOB",
    "Address": "ADDR1",
    "Address_2": "ADDR2",
    "ContactEmail": "EMAIL",
    "LMS_Account": "LMS",
    "PathNo": "PATH",
    "WardofState": "WARD",
    "Sector": "SECTOR",
    "Parish": "PARISH",
    "Region": "REGION",
    "Prefcode1": "PCODE1",
    "First_Preference": "PNAME1",
    "Prefcode2": "PCODE2",
    "Second_Preference": "PNAME2",
    "Prefcode3": "PCODE3",
    "Third_Preference": "PNAME3",
    "Prefcode4": "PCODE4",
    "Fourth_Preference": "PNAME4",
    "Prefcode5": "PCODE5",
    "Fifth_Preference": "PNAME5",
    "Prefcode6": "CCODE1",
    "Sixth_Preference": "CNAME1",
    "Prefcode7": "CCODE2",
    "Seventh_Preference": "CNAME2",
    "SRN": "SRN"
}

paper_names = [
    "Ability",  # 01
    "Mathematics PT",  # 02
    "Language Arts PT",  # 03
    "Mathematics CBT",  # 04
    "Science CBT",  # 05
    "Social Studies CBT",  # 06
    "Language Arts CBT"  # 07
]

# Load files
df_data = pd.read_excel(data_path)
wb_input = load_workbook(input_path)
ws = wb_input.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

rows_to_add = []

for _, row in df_data.iterrows():
    new_row = []
    for header in headers:
        match = next((src for src, dest in field_map.items() if dest == header), None)

        if match in row:
            val = row[match]
            if header in ["PATH"]:
                val = "Y" if pd.notna(val) and str(val).strip() != "" else "N"
            new_row.append(val)

        elif header == "TEL_NUM":
            if pd.notna(row.get("MotherContact")) and str(row.get("MotherContact")).strip():
                val = row.get("MotherContact", "")
            elif pd.notna(row.get("FatherContact")) and str(row.get("FatherContact")).strip():
                val = row.get("FatherContact", "")
            elif pd.notna(row.get("GuardianContact")) and str(row.get("GuardianContact")).strip():
                val = row.get("GuardianContact", "")
            else:
                val = ""
            new_row.append(val)

        elif header == "GNAME":
            if pd.notna(row.get("MotherContact")) and str(row.get("MotherContact")).strip():
                val = row.get("Mother", "")
            elif pd.notna(row.get("FatherContact")) and str(row.get("FatherContact")).strip():
                val = row.get("Father", "")
            elif pd.notna(row.get("GuardianContact")) and str(row.get("GuardianContact")).strip():
                val = row.get("Gaurdian", "")
            elif pd.notna(row.get("Mother")) and str(row.get("Mother")).strip():
                val = row.get("Mother", "")
            elif pd.notna(row.get("Father")) and str(row.get("Father")).strip():
                val = row.get("Father", "")
            elif pd.notna(row.get("Gaurdian")) and str(row.get("Gaurdian")).strip():
                val = row.get("Gaurdian", "")
            new_row.append(val)

        elif header.startswith("PAPER0") and header[-2:].isdigit():
            index = int(header[-2:]) - 1
            val = paper_names[index] if index < len(paper_names) else ""
            new_row.append(val)
        else:
            new_row.append(None)

    rows_to_add.append(new_row)

# Append rows
for row in rows_to_add:
    ws.append(row)

# Save file
wb_input.save(output_path)
messagebox.showinfo("Success", f"Data successfully saved to:\n{output_path}")
