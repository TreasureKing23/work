import customtkinter as ctk
from tkinter import filedialog
from openpyxl import load_workbook
import pandas as pd
import threading
import os

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

class ExcelMapperApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Excel Data Formatter")
        self.geometry("700x550")
        self.resizable(False, False)

        # paths
        self.input_path = ""
        self.data_path = ""
        self.output_path = ""

        # exam selector
        self.selected_exam = ctk.StringVar(value="PEP6")

        self._build_ui()

    # ------------------------------------------------------------------ UI
    def _build_ui(self):
        ctk.CTkLabel(self, text="📄 Excel Data Formatter", font=ctk.CTkFont(size=24, weight="bold")).pack(pady=20)

        ctk.CTkLabel(self, text="Select Exam:").pack()
        ctk.CTkOptionMenu(self, variable=self.selected_exam, values=["PEP4", "PEP5", "PEP6"]).pack(pady=5)

        self.input_btn = ctk.CTkButton(self, text="Select Input Specification", command=self._select_input)
        self.input_btn.pack(pady=10)

        self.data_btn = ctk.CTkButton(self, text="Select Data File", command=self._select_data)
        self.data_btn.pack(pady=10)

        self.output_btn = ctk.CTkButton(self, text="Choose Output Location", command=self._select_output)
        self.output_btn.pack(pady=10)

        self.run_btn = ctk.CTkButton(self, text="Run Formatting", command=self._start_processing)
        self.run_btn.pack(pady=20)

        self.progress = ctk.CTkProgressBar(self, width=400)
        self.progress.set(0)
        self.progress.pack(pady=10)

        self.status_label = ctk.CTkLabel(self, text="")
        self.status_label.pack()

    # ------------------------------------------------------------------ Browsers
    def _select_input(self):
        path = filedialog.askopenfilename(title="Select Input Specification", filetypes=[("Excel", "*.xlsx")])
        if path:
            self.input_path = path
            self.input_btn.configure(text=os.path.basename(path))

    def _select_data(self):
        path = filedialog.askopenfilename(title="Select Data File", filetypes=[("Excel", "*.xlsx")])
        if path:
            self.data_path = path
            self.data_btn.configure(text=os.path.basename(path))

    def _select_output(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Save Output As", filetypes=[("Excel", "*.xlsx")])
        if path:
            self.output_path = path
            self.output_btn.configure(text=os.path.basename(path))

    # ------------------------------------------------------------------ Thread wrapper
    def _start_processing(self):
        threading.Thread(target=self._run_mapping, daemon=True).start()

    # ------------------------------------------------------------------ Core processing
    def _run_mapping(self):
        if not all([self.input_path, self.data_path, self.output_path]):
            self.status_label.configure(text="⚠️ Please select all files.")
            return

        self.status_label.configure(text="🔄 Processing …")
        self.progress.set(0)
        self.update_idletasks()

        # ------------- mapping dict & paper names
        field_map = {
            "ExamCode": "EXAMID", "ExamYear": "PERIODID", "EXCID": "CANID", "ERN": "ERN",
            "FirstName": "FNAME", "Lastname": "SURNAME", "Middlename": "MNAME",
            "SchoolId": "SCHOOLID", "SchoolName": "SCHOOLN", "Gender": "GENDER",
            "DateofBirth": "DOB", "Address": "ADDR1", "Address_2": "ADDR2",
            "ContactEmail": "EMAIL", "LMS_Account": "LMS", "PathNo": "PATH",
            "WardofState": "WARD", "Sector": "SECTOR", "Parish": "PARISH", "Region": "REGION",
            "Prefcode1": "PCODE1", "First_Preference": "PNAME1", "Prefcode2": "PCODE2",
            "Second_Preference": "PNAME2", "Prefcode3": "PCODE3", "Third_Preference": "PNAME3",
            "Prefcode4": "PCODE4", "Fourth_Preference": "PNAME4", "Prefcode5": "PCODE5",
            "Fifth_Preference": "PNAME5", "Prefcode6": "CCODE1", "Sixth_Preference": "CNAME1",
            "Prefcode7": "CCODE2", "Seventh_Preference": "CNAME2", "SRN": "SRN"
        }

        papers = {
            "PEP6": [
                "Ability", "Mathematics PT", "Language Arts PT",
                "Mathematics CBT", "Science CBT", "Social Studies CBT", "Language Arts CBT"
            ],
            "PEP5": [
                "Mathematics PT", "Science PT", "Social Studies PT", "Language PT"
            ],
            "PEP4": [
                "Mathematics PT", "Numeracy", "Language PT", "Literacy"
            ]
        }

        # ------------- load files
        df_src = pd.read_excel(self.data_path)
        wb_template = load_workbook(self.input_path)
        ws = wb_template.active
        headers = [cell.value for cell in ws[1]]

        paper_names = papers.get(self.selected_exam.get(), [])

        rows_to_append = []
        for r_idx, src_row in df_src.iterrows():
            new_row = []
            for header in headers:
                src_key = next((k for k, v in field_map.items() if v == header), None)

                if src_key in src_row:
                    val = src_row[src_key]
                    if header in ("PATH", "WARD"):
                        val = "Y" if pd.notna(val) and str(val).strip() else "N"
                    new_row.append(val)

                elif header == "TEL_NUM":
                    val = src_row.get("MotherContact") or src_row.get("FatherContact") or src_row.get("GuardianContact") or ""
                    new_row.append(val)

                elif header == "GNAME":
                    if src_row.get("MotherContact"):
                        val = src_row.get("Mother", "")
                    elif src_row.get("FatherContact"):
                        val = src_row.get("Father", "")
                    elif src_row.get("GuardianContact"):
                        val = src_row.get("Gaurdian", "")
                    else:
                        val = src_row.get("Mother") or src_row.get("Father") or src_row.get("Gaurdian") or ""
                    new_row.append(val)

                elif header.startswith("PAPER0") and header[-2:].isdigit():
                    idx = int(header[-2:]) - 1
                    new_row.append(paper_names[idx] if idx < len(paper_names) else "")
                else:
                    new_row.append(None)

            rows_to_append.append(new_row)
            self.progress.set((r_idx + 1) / len(df_src))
            self.update_idletasks()

        # ------------- write to template workbook
        for row in rows_to_append:
            ws.append(row)

        wb_template.save(self.output_path)

        # ------------- create fixed‑width text file -------------------
        txt_path = self.output_path.replace(".xlsx", "_fixed.txt")
        df_final = pd.DataFrame(rows_to_append, columns=headers)

        # compute widths
        col_widths = [max(len(str(col)), df_final[col].astype(str).map(len).max()) for col in df_final.columns]

        with open(txt_path, "w", encoding="utf-8") as f:
            # header row
            f.write(''.join(str(col).ljust(w + 2) for col, w in zip(df_final.columns, col_widths)) + "\n")
            # data rows
            for _, row in df_final.iterrows():
                f.write(''.join(str(cell).ljust(w + 2) for cell, w in zip(row, col_widths)) + "\n")

        self.progress.set(1)
        self.status_label.configure(text=f"✅ Complete! Saved:\n{os.path.basename(self.output_path)} and {os.path.basename(txt_path)}")

if __name__ == "__main__":
    app = ExcelMapperApp()
    app.mainloop()
