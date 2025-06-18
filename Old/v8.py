import customtkinter as ctk
from tkinter import filedialog
from openpyxl import load_workbook
import pandas as pd
import threading

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

class ExcelMapperApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Excel Data Formatter")
        self.geometry("700x550")
        self.resizable(False, False)

        self.input_path = ""
        self.data_path = ""
        self.output_path = ""

        self.selected_exam = ctk.StringVar(value="PEP6")

        self.build_ui()

    def build_ui(self):
        ctk.CTkLabel(self, text="📄 Excel Data Formatter", font=ctk.CTkFont(size=24, weight="bold")).pack(pady=20)

        ctk.CTkLabel(self, text="Select Exam:").pack()
        self.exam_dropdown = ctk.CTkOptionMenu(self, variable=self.selected_exam, values=["PEP4", "PEP5", "PEP6"])
        self.exam_dropdown.pack(pady=5)

        self.input_btn = ctk.CTkButton(self, text="Select Input Specification", command=self.select_input)
        self.input_btn.pack(pady=10)

        self.data_btn = ctk.CTkButton(self, text="Select Data File", command=self.select_data)
        self.data_btn.pack(pady=10)

        self.output_btn = ctk.CTkButton(self, text="Choose Output Location", command=self.select_output)
        self.output_btn.pack(pady=10)

        self.run_btn = ctk.CTkButton(self, text="Run Formatting", command=self.start_mapping)
        self.run_btn.pack(pady=20)

        self.progress = ctk.CTkProgressBar(self, width=400)
        self.progress.set(0)
        self.progress.pack(pady=10)

        self.status_label = ctk.CTkLabel(self, text="")
        self.status_label.pack()

    def select_input(self):
        self.input_path = filedialog.askopenfilename(title="Select Input Specification")
        self.input_btn.configure(text="Input Specification Selected")

    def select_data(self):
        self.data_path = filedialog.askopenfilename(title="Select Data File")
        self.data_btn.configure(text="Data File Selected")

    def select_output(self):
        self.output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Save Output As")
        self.output_btn.configure(text="Output Location Set")

    def start_mapping(self):
        threading.Thread(target=self.run_mapping).start()

    def run_mapping(self):
        if not self.input_path or not self.data_path or not self.output_path:
            self.status_label.configure(text="⚠️ Please select all files.")
            return

        self.progress.set(0)
        self.status_label.configure(text="🔄 Processing...")
        self.update_idletasks()

        field_map = {
            "ExamCode": "EXAMID", "ExamYear": "PERIODID", "EXCID": "CANID", "ERN": "ERN",
            "FirstName": "FNAME", "Lastname": "SURNAME", "Middlename": "MNAME",
            "SchoolId": "SCHOOLID", "SchoolName": "SCHOOLN", "Gender": "GENDER",
            "DateofBirth": "DOB", "Address": "ADDR1", "Address_2": "ADDR2",
            "ContactEmail": "EMAIL", "LMS_Account": "LMS", "PathNo": "PATH",
            "WardofState": "WARD", "Sector": "SECTOR", "Parish": "PARISH", "Region": "REGION",
            "Prefcode1": "PCODE1", "First_Preference": "PNAME1", "Prefcode2": "PCODE2", 
            "Second_Preference": "PNAME2", "Prefcode3": "PCODE3", "Third_Preference": "PNAME3",
            "Prefcode4": "PCODE4", "Fourth_Preference": "PNAME4", "Prefcode5": "PCODE5",
            "Fifth_Preference": "PNAME5", "Prefcode6": "CCODE1", "Sixth_Preference": "CNAME1",
            "Prefcode7": "CCODE2", "Seventh_Preference": "CNAME2", "SRN": "SRN"
        }

        PEP6paper_names = [
            "Ability", "Mathematics PT", "Language Arts PT",
            "Mathematics CBT", "Science CBT", "Social Studies CBT", "Language Arts CBT"
        ]

        PEP5paper_names = [
            "Mathematics PT", "Science PT", "Social Studies PT", "Language PT"
        ]

        PEP4paper_names = [
            "Mathematics PT", "Numeracy", "Language PT",  "Literacy"
        ]

        df = pd.read_excel(self.data_path)
        wb = load_workbook(self.input_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        selected_exam = self.selected_exam.get()
        if selected_exam == "PEP5":
            paper_names = PEP5paper_names
        elif selected_exam == "PEP6":
            paper_names = PEP6paper_names
        elif selected_exam == "PEP4":
            paper_names = PEP4paper_names
        else:
            paper_names = []

        rows_to_add = []
        for idx, row in df.iterrows():
            new_row = []
            for header in headers:
                key = next((k for k, v in field_map.items() if v == header), None)

                if key in row:
                    val = row[key]
                    if header == "PATH":
                        val = "Y" if pd.notna(val) and str(val).strip() else "N"
                    elif header == "WARD":
                        val = "Y" if pd.notna(row.get("WardofState")) else "N"
                    new_row.append(val)

                elif header == "TEL_NUM":
                    if pd.notna(row.get("MotherContact")) and str(row.get("MotherContact")).strip():
                        val = row.get("MotherContact", "")
                    elif pd.notna(row.get("FatherContact")) and str(row.get("FatherContact")).strip():
                        val = row.get("FatherContact", "")
                    elif pd.notna(row.get("GuardianContact")) and str(row.get("GuardianContact")).strip():
                        val = row.get("GuardianContact", "")
                    else:
                        val = ""
                    new_row.append(val)

                elif header == "GNAME":
                    if pd.notna(row.get("MotherContact")) and str(row.get("MotherContact")).strip():
                        val = row.get("Mother", "")
                    elif pd.notna(row.get("FatherContact")) and str(row.get("FatherContact")).strip():
                        val = row.get("Father", "")
                    elif pd.notna(row.get("GuardianContact")) and str(row.get("GuardianContact")).strip():
                        val = row.get("Gaurdian", "")
                    elif pd.notna(row.get("Mother")) and str(row.get("Mother")).strip():
                        val = row.get("Mother", "")
                    elif pd.notna(row.get("Father")) and str(row.get("Father")).strip():
                        val = row.get("Father", "")
                    elif pd.notna(row.get("Gaurdian")) and str(row.get("Gaurdian")).strip():
                        val = row.get("Gaurdian", "")
                    else:
                        val = ""
                    new_row.append(val)

                elif header.startswith("PAPER0") and header[-2:].isdigit():
                    index = int(header[-2:]) - 1
                    new_row.append(paper_names[index] if index < len(paper_names) else "")
                else:
                    new_row.append(None)
            rows_to_add.append(new_row)
            self.progress.set((idx + 1) / len(df))
            self.update_idletasks()

        for row in rows_to_add:
            ws.append(row)

        wb.save(self.output_path)

       

        self.progress.set(1)
        self.status_label.configure(text="✅ Formatting Complete!")

if __name__ == "__main__":
    app = ExcelMapperApp()
    app.mainloop()
