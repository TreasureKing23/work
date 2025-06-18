import pandas as pd
from openpyxl import load_workbook

# File paths
dwb_path = "Data.xlsx"
iwb_path = "Input.xlsx"
output_path="Updated.xlsx"

# Load Data 
dwb = pd.read_excel(dwb_path)

# Mapping from Data -> Input
field_map = {
    "ExamCode": "EXAMID",
    "ExamYear": "PERIODID",
    "EXCID":"CANID",
    "ERN": "ERN",
    "FirstName": "FNAME",
    "Lastname": "SURNAME",
    "Middlename": "MNAME",
    "SchoolId": "SCHOOLID",
    "SchoolName": "SCHOOLN",
    "Gender": "GENDER",
    "DateofBirth": "DOB",
    "Address": "ADDR1",
    "Address_2": "ADDR2",
    "ContactEmail": "EMAIL",
    "LMS_Account": "LMS",
    "PathNo": "PATH",
    "WardofState": "WARD",
    "Sector": "SECTOR",
    "Parish": "PARISH",
    "Region": "REGION",
    "Prefcode1": "PCODE1",
    "First_Preference": "PNAME1",
    "Prefcode2": "PCODE2",
    "Second_Preference": "PNAME2",
    "Prefcode3": "PCODE3",
    "Third_Preference": "PNAME3",
    "Prefcode4": "PCODE4",
    "Fourth_Preference": "PNAME4",
    "Prefcode5": "PCODE5",
    "Fifth_Preference": "PNAME5",
    "Prefcode6": "CCODE1",
    "Sixth_Preference":"CNAME1",
    "Prefcode7": "CCODE2",
    "Seventh_Preference":"CNAME2",
    "SRN": "SRN"
}

paper_names=[
    "Ability", #01
    "Mathematics PT", #02
    "Language Arts PT", #03
    "Mathematics CBT", #04
    "Science CBT", #05
    "Social Studies CBT",#06
    "Language Arts CBT" #07
]

#load

df_dwb = pd.read_excel(dwb_path)
iwb= load_workbook(iwb_path)
ws1 = iwb.active

headers = [cell.value for cell in next(ws1.iter_rows(min_row=1,max_row=1))]

rows_to_add =[]

for _, row in df_dwb.iterrows():
    new_row=[]
    for header in headers:
        match = next((src for src, dest in field_map.items() if dest == header), None)

        if match in row:
            val = row[match]


            if header in ["PATH"]:
                val = "Y" if pd.notna(val) and str(val).strip() != "" else "N"
            new_row.append(val)


        elif header =="TEL_NUM":
            if pd.notna(row.get("MotherContact")):
                val= row.get("MotherContact","")
            elif pd.notna(row.get("FatherContact")):
                val= row.get("FatherContact","")
            elif pd.notna(row.get("GuardianContact")):
                val= row.get("GuardianContact","")
            else:
                val = ""
            new_row.append(val)


        elif header =="GNAME":
            if pd.notna(row.get("MotherContact")):
                val= row.get("Mother","")
            elif pd.notna(row.get("FatherContact")):
                val= row.get("Father","")
            elif pd.notna(row.get("GuardianContact")):
                val= row.get("Gaurdian","")
            else:
                val = ""
            new_row.append(val)


        elif header.startswith("PAPER0") and header[-2:].isdigit():
            index = int(header[-2:]) - 1
            val = paper_names[index] if index < len(paper_names) else ""
            new_row.append(val)
        else:
            new_row.append(None)

    rows_to_add.append(new_row)

for row in rows_to_add:
    ws1.append(row)

iwb.save(output_path)
print(f"done saved to {output_path}")