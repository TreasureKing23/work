import pandas as pd
from openpyxl import load_workbook
from tkinter import Tk, Label, Entry, Button, filedialog, messagebox
from tkinter.ttk import Progressbar

class ExcelImporter:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Import Tool")
        self.root.geometry("800x300")

        Label(root, text="Input Excel File (Workbook1)").grid(row=0, column=0, padx=10, pady=10, sticky='w')
        self.input_entry = Entry(root, width=50)
        self.input_entry.grid(row=0, column=1)
        Button(root, text="Browse", command=self.browse_input).grid(row=0, column=2)

        Label(root, text="Data Excel File (Workbook2)").grid(row=1, column=0, padx=10, pady=10, sticky='w')
        self.data_entry = Entry(root, width=50)
        self.data_entry.grid(row=1, column=1)
        Button(root, text="Browse", command=self.browse_data).grid(row=1, column=2)

        Label(root, text="Output File Name").grid(row=2, column=0, padx=10, pady=10, sticky='w')
        self.output_entry = Entry(root, width=50)
        self.output_entry.grid(row=2, column=1)
        Button(root, text="Browse", command=self.save_output).grid(row=2, column=2)

        self.pb = Progressbar(root, length=500, mode='determinate')
        self.pb.grid(row=3, column=0, columnspan=3, pady=20)

        Button(root, text="Start Import", command=self.start_import).grid(row=4, column=1, pady=10)

    def browse_input(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if path:
            self.input_entry.delete(0, 'end')
            self.input_entry.insert(0, path)

    def browse_data(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if path:
            self.data_entry.delete(0, 'end')
            self.data_entry.insert(0, path)

    def save_output(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if path:
            self.output_entry.delete(0, 'end')
            self.output_entry.insert(0, path)

    def start_import(self):
        input_path = self.input_entry.get()
        data_path = self.data_entry.get()
        output_path = self.output_entry.get()

        if not input_path or not data_path or not output_path:
            messagebox.showerror("Missing Input", "All fields must be filled.")
            return

        try:
            self.run_import(input_path, data_path, output_path)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def run_import(self, input_path, data_path, output_path):
        field_map = {
            "ExamCode": "EXAMID", "ExamYear": "PERIODID", "EXCID": "CANID", "ERN": "ERN",
            "FirstName": "FNAME", "Lastname": "SURNAME", "Middlename": "MNAME",
            "SchoolId": "SCHOOLID", "SchoolName": "SCHOOLN", "Gender": "GENDER",
            "DateofBirth": "DOB", "Address": "ADDR1", "Address_2": "ADDR2",
            "ContactEmail": "EMAIL", "LMS_Account": "LMS", "PathNo": "PATH", "WardofState": "WARD",
            "Sector": "SECTOR", "Parish": "PARISH", "Region": "REGION",
            "Prefcode1": "PCODE1", "First_Preference": "PNAME1",
            "Prefcode2": "PCODE2", "Second_Preference": "PNAME2",
            "Prefcode3": "PCODE3", "Third_Preference": "PNAME3",
            "Prefcode4": "PCODE4", "Fourth_Preference": "PNAME4",
            "Prefcode5": "PCODE5", "Fifth_Preference": "PNAME5",
            "Prefcode6": "CCODE1", "Sixth_Preference": "CNAME1",
            "Prefcode7": "CCODE2", "Seventh_Preference": "CNAME2", "SRN": "SRN"
        }

        paper_names = [
            "Ability", "Mathematics PT", "Language Arts PT",
            "Mathematics CBT", "Science CBT", "Social Studies CBT", "Language Arts CBT"
        ]

        df_data = pd.read_excel(data_path)
        wb_input = load_workbook(input_path)
        ws = wb_input.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        self.pb["maximum"] = len(df_data)

        for i, (_, row) in enumerate(df_data.iterrows()):
            new_row = []
            for header in headers:
                match = next((src for src, dest in field_map.items() if dest == header), None)

                if match in row:
                    val = row[match]
                    if header in ["PATH", "WARD"]:
                        val = "Y" if pd.notna(val) and str(val).strip() else "N"
                    new_row.append(val)

                elif header == "TEL_NIM":
                    for field in ["MotherContact", "FatherContact", "GuardianContact"]:
                        val = row.get(field, "")
                        if pd.notna(val) and str(val).strip():
                            break
                    new_row.append(val)

                elif header == "GNAME":
                    for contact, name in [("MotherContact", "Mother"), ("FatherContact", "Father"), ("GuardianContact", "Gaurdian")]:
                        if pd.notna(row.get(contact)) and str(row.get(contact)).strip():
                            val = row.get(name, "")
                            break
                    else:
                        val = ""
                    new_row.append(val)

                elif header.startswith("PAPER0") and header[-2:].isdigit():
                    index = int(header[-2:]) - 1
                    val = paper_names[index] if index < len(paper_names) else ""
                    new_row.append(val)

                else:
                    new_row.append(None)

            ws.append(new_row)
            self.pb["value"] = i + 1
            self.root.update()

        wb_input.save(output_path)
        messagebox.showinfo("Success", f"Data saved to {output_path}")

if __name__ == "__main__":
    root = Tk()
    app = ExcelImporter(root)
    root.mainloop()
